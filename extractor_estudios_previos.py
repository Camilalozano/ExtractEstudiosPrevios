import io
import re
import sys
import zipfile
import unicodedata
from pathlib import Path
from typing import Dict, List

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

OUTPUT_COLUMNS = [
    "archivo_pdf",
    "Fecha de Elaboración",
    "OBJETO",
    "TIPO DE CONTRATO",
    "PLAZO DE EJECUCIÓN",
    "VALOR DEL CONTRATO",
    "FORMA DE PAGO",
    "OBLIGACIONES ESPECÍFICAS DEL CONTRATISTA",
    "texto_extraido_chars",
    "extraccion_ok",
]

FOOTER_BLOCK_PATTERNS = [
    r"\n?\s*\d+\s*\n\s*Formato de Estudios y Documentos Previos\s*[–-]\s*\n\s*Contrataci[oó]n Directa\s*\n\s*C[ÓO]DIGO:\s*F1_P11_C\s*\n\s*VERSI[ÓO]N:?\s*3\s*\n\s*Proceso Gesti[oó]n Contractual\s*\n\s*FECHA DE APROBACI[ÓO]N:?\s*\n\s*20/06/2024\s*\n\s*P[aá]gina\s*\d+\s*de\s*\d+\s*\n\s*Piensa en el medio ambiente antes de imprimir este documento\.\s*\n\s*Cualquier copia impresa de este documento se considera como COPIA NO CONTROLADA\s*\n\s*LOS DATOS PROPORCIONADOS SER[ÁA]N TRATADOS.*?https://agenciaatenea\.gov\.co/?\s*",
    r"\n?\s*Piensa en el medio ambiente antes de imprimir este documento\.\s*\n\s*Cualquier copia impresa de este documento se considera como COPIA NO CONTROLADA\s*\n\s*LOS DATOS PROPORCIONADOS SER[ÁA]N TRATADOS.*?https://agenciaatenea\.gov\.co/?\s*",
]

FOOTER_LINE_PATTERNS = [
    r"^\s*\d+\s*$",
    r"^\s*Formato de Estudios y Documentos Previos\s*[–-]\s*$",
    r"^\s*Contrataci[oó]n Directa\s*$",
    r"^\s*C[ÓO]DIGO:\s*F1_P11_C\s*$",
    r"^\s*VERSI[ÓO]N:?\s*3\s*$",
    r"^\s*Proceso Gesti[oó]n Contractual\s*$",
    r"^\s*FECHA DE APROBACI[ÓO]N:?\s*$",
    r"^\s*20/06/2024\s*$",
    r"^\s*P[aá]gina\s*\d+\s*de\s*\d+\s*$",
    r"^\s*Piensa en el medio ambiente antes de imprimir este documento\.\s*$",
    r"^\s*Cualquier copia impresa de este documento se considera como COPIA NO CONTROLADA\s*$",
    r"^\s*LOS DATOS PROPORCIONADOS SER[ÁA]N TRATADOS.*$",
    r"^\s*AGENCIA PUBLICDA EN LA P[ÁA]GINA WEB.*$",
]


def normalize_spaces(text: str) -> str:
    text = text or ""
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()



def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    cleaned = []
    for ch in text:
        cat = unicodedata.category(ch)
        if cat.startswith("C") and ch not in ["\n", "\t"]:
            continue
        cleaned.append(ch)
    return normalize_spaces("".join(cleaned))



def clean_section_text(text: str) -> str:
    if not text:
        return ""
    cleaned = text.replace("\r\n", "\n").replace("\r", "\n")
    for pattern in FOOTER_BLOCK_PATTERNS:
        cleaned = re.sub(pattern, "\n", cleaned, flags=re.IGNORECASE | re.DOTALL)

    kept_lines = []
    for line in cleaned.split("\n"):
        if any(re.match(pat, line, flags=re.IGNORECASE) for pat in FOOTER_LINE_PATTERNS):
            continue
        kept_lines.append(line)

    cleaned = "\n".join(kept_lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    return cleaned.strip()



def normalize_inline(text: str) -> str:
    text = clean_section_text(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .;:\n\t")



def extract_text_with_pymupdf(pdf_bytes: bytes) -> str:
    if fitz is None:
        return ""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        return "\n".join((page.get_text("text") or "") for page in doc)
    except Exception:
        return ""



def extract_text_with_pypdf(pdf_bytes: bytes) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        pages = []
        for page in reader.pages:
            try:
                pages.append(page.extract_text() or "")
            except Exception:
                pages.append("")
        return "\n".join(pages)
    except Exception:
        return ""



def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    text = extract_text_with_pymupdf(pdf_bytes)
    if len(normalize_text(text)) >= 200:
        return text
    alt = extract_text_with_pypdf(pdf_bytes)
    if len(normalize_text(alt)) > len(normalize_text(text)):
        return alt
    return text




SECTION_2_HEADING_RE = re.compile(
    r"(?im)^\s*2\s*[.\-]\s*(?:ESPECIFICACIONES|DESCRIPCI[ÓO]N)\s+DEL\s+OBJETO\s+A\s+CONTRATAR\b.*$"
)

SECTION_2_END_RE = re.compile(
    r"(?im)^\s*3\s*[.\-]\s*(?:MODALIDAD|FUNDAMENTOS|JUSTIFICACI[ÓO]N|CRITERIOS|AN[ÁA]LISIS|ESTUDIO|GARANT[ÍI]AS)\b"
)

OBJETO_LABEL_RE = re.compile(
    r"(?im)^\s*(?:[a-z0-9]+\s*[\.\)]\s*)?OBJETO\s*:?\s*(?!A\s+CONTRATAR\b|,)(.*)$"
)

NEXT_OBJECT_FIELD_RE = re.compile(
    r"(?im)^\s*(?:[a-z0-9]+\s*[\.\)]\s*)?"
    r"(?:TIPO\s+DE\s+CONTRATO|CODIFICACI[ÓO]N\s+CLASIFICADOR|CLASIFICADOR\s+DE\s+BIENES\s+Y\s+SERVICIOS|"
    r"PLAZO\s+DE\s+EJECUCI[ÓO]N|VALOR\s+DEL\s+CONTRATO|CERTIFICADO\s+DE\s+DISPONIBILIDAD\s+PRESUPUESTAL|"
    r"FORMA\s+DE\s+PAGO|OBLIGACIONES\s+GENERALES|OBLIGACIONES\s+ESPEC[ÍI]FICAS|SUPERVISI[ÓO]N|GARANT[ÍI]AS)\s*:?"
)

CONTRACT_OBJECT_VERB_RE = re.compile(
    r"\b(?:Prestar|Contratar|Adquirir|Suministrar|Realizar|Brindar|Apoyar|Desarrollar|Ejecutar|Implementar|"
    r"Fortalecer|Aunar|Celebrar)\b",
    re.IGNORECASE,
)


def get_specs_section(text: str) -> str:
    """
    Ubica el bloque 2 de los Estudios Previos.

    Mejora frente a la regla anterior:
    - Acepta "2. ESPECIFICACIONES DEL OBJETO A CONTRATAR".
    - Acepta "2. DESCRIPCIÓN DEL OBJETO A CONTRATAR".
    - Corta al inicio del numeral 3, aunque el título del numeral 3 tenga variaciones.
    - Elimina encabezados/pies de página antes de buscar los campos.
    """
    text = clean_section_text(text)
    matches = list(SECTION_2_HEADING_RE.finditer(text))
    if not matches:
        return text

    chosen = matches[0]
    for match in matches:
        window = text[match.end(): match.end() + 8000]
        if OBJETO_LABEL_RE.search(window):
            chosen = match
            break

    start = chosen.end()
    end_match = SECTION_2_END_RE.search(text, start)
    end = end_match.start() if end_match else len(text)
    return clean_section_text(text[start:end])



def clean_objeto_text(block: str) -> str:
    """
    Limpieza específica del campo OBJETO.
    Evita que queden códigos, encabezados o fragmentos ajenos antes del verbo contractual.
    """
    block = clean_section_text(block)
    block = re.sub(r"^(?:[a-z0-9]+\s*[\.\)]\s*)?OBJETO\s*:?\s*", "", block, flags=re.IGNORECASE).strip()

    verb_match = CONTRACT_OBJECT_VERB_RE.search(block)
    if verb_match and 0 < verb_match.start() < 120:
        prefix = block[:verb_match.start()]
        # Si antes del verbo solo hay códigos, consecutivos, separadores o texto basura de extracción, se elimina.
        has_meaningful_words = bool(re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]{4,}\s+[A-Za-zÁÉÍÓÚÑáéíóúñ]{4,}", prefix))
        has_code_noise = bool(re.search(r"\d|_|-|/", prefix))
        if has_code_noise or not has_meaningful_words:
            block = block[verb_match.start():]

    block = re.sub(r"\s+", " ", block)
    return block.strip(" .;:\n\t")



def extract_objeto(text: str) -> str:
    """
    Extrae el OBJETO contractual dentro del numeral 2, usando como ancla el campo real
    "OBJETO" y no el título "ESPECIFICACIONES DEL OBJETO A CONTRATAR".

    Regla:
    1. Buscar el numeral 2 de especificaciones/descripción del objeto.
    2. Dentro de ese bloque, buscar una línea tipo:
       - a. OBJETO:
       - 1. OBJETO:
       - OBJETO:
       - OBJETO
    3. Tomar el texto posterior hasta el siguiente campo estructural:
       TIPO DE CONTRATO, CODIFICACIÓN, PLAZO, VALOR, FORMA DE PAGO, OBLIGACIONES, etc.
    4. Limpiar pies de página, saltos y códigos residuales.
    """
    section = get_specs_section(text)
    candidates = []

    for match in OBJETO_LABEL_RE.finditer(section):
        start = match.start(1) if match.group(1).strip() else match.end()
        next_field = NEXT_OBJECT_FIELD_RE.search(section, start)
        end = next_field.start() if next_field else len(section)

        candidate = clean_objeto_text(section[start:end])
        if candidate:
            candidates.append(candidate)

    if not candidates:
        return ""

    def score_candidate(candidate: str) -> float:
        score = 0.0
        if CONTRACT_OBJECT_VERB_RE.match(candidate):
            score += 100
        if 40 <= len(candidate) <= 1200:
            score += 50
        # Se favorecen textos de longitud razonable; se penalizan textos demasiado largos o demasiado cortos.
        score -= abs(len(candidate) - 250) / 100
        return score

    return max(candidates, key=score_candidate)



def extract_fecha_elaboracion(text: str) -> str:
    patterns = [
        r"Fecha\s+de\s+Elaboraci[oó]n\s*[:\-]?\s*\n\s*([0-3]?\d[/-][0-1]?\d[/-](?:19|20)\d{2})",
        r"Fecha\s+de\s+Elaboraci[oó]n\s*[:\-]?\s*\n\s*([A-Za-zÁÉÍÓÚÑáéíóúñ]+\s+\d{1,2}\s+de\s+(?:19|20)\d{2})",
        r"Fecha\s+de\s+Elaboraci[oó]n\s*[:\-]?\s*\n\s*((?:19|20)\d{2})\b",
        r"Fecha\s+de\s+Elaboraci[oó]n\s*[:\-]?\s*([0-3]?\d[/-][0-1]?\d[/-](?:19|20)\d{2})",
        r"Fecha\s+de\s+Elaboraci[oó]n\s*[:\-]?\s*((?:19|20)\d{2})\b",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return normalize_inline(m.group(1))
    return ""



def extract_tipo_contrato(text: str) -> str:
    text = get_specs_section(text)
    m = re.search(
        r"TIPO DE CONTRATO\s*:?[\s]*(.+?)(?=\n\s*(?:[a-z]\.\s*)?(?:CODIFICACI[ÓO]N CLASIFICADOR BIENES Y SERVICIOS|PLAZO DE EJECUCI[ÓO]N)\s*:?)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    return normalize_inline(m.group(1)) if m else ""



def extract_plazo(text: str) -> str:
    text = get_specs_section(text)
    m = re.search(
        r"PLAZO DE EJECUCI[ÓO]N\s*:?[\s]*(.+?)(?=\n\s*(?:[a-z]\.\s*)?VALOR DEL CONTRATO\s*:?)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    return normalize_inline(m.group(1)) if m else ""



def extract_valor(text: str) -> str:
    text = get_specs_section(text)
    m = re.search(
        r"VALOR DEL CONTRATO\s*:?[\s]*(.+?)(?=\n\s*(?:[a-z]\.\s*)?(?:CERTIFICADO DE DISPONIBILIDAD PRESUPUESTAL|FORMA DE PAGO)\s*:?)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    block = clean_section_text(m.group(1)) if m else ""
    block = re.split(r"(?:^|\n)\s*Nota\s*\d+\s*:", block, maxsplit=1, flags=re.IGNORECASE)[0]
    return normalize_inline(block)



def extract_forma_pago(text: str) -> str:
    text = get_specs_section(text)
    m = re.search(
        r"FORMA DE PAGO\s*:?[\s]*(.+?)(?=\n\s*(?:[a-z]\.\s*)?OBLIGACIONES GENERALES DEL CONTRATISTA|\n\s*3\.\s*MODALIDAD DE SELECCI[ÓO]N Y JUSTIFICACI[ÓO]N DE LA MISMA)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    return clean_section_text(m.group(1)) if m else ""



def extract_obligaciones(text: str) -> str:
    m = re.search(
        r"OBLIGACIONES ESPEC[ÍI]FICAS DEL CONTRATISTA\s*(.+?)(?=\n\s*3\.\s*MODALIDAD DE SELECCI[ÓO]N Y JUSTIFICACI[ÓO]N DE LA MISMA|\n\s*4\.\s*CRITERIOS PARA SELECCIONAR)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    block = clean_section_text(m.group(1)) if m else ""
    block = re.sub(
        r"^\s*Además de las obligaciones generales, incluidas.*?las siguientes obligaciones:\s*",
        "",
        block,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return block.strip()



def build_row_from_pdf(pdf_bytes: bytes, filename: str) -> Dict[str, object]:
    raw_text = extract_text_from_pdf_bytes(pdf_bytes)
    text = normalize_text(raw_text)

    row = {
        "archivo_pdf": Path(filename).name,
        "Fecha de Elaboración": extract_fecha_elaboracion(text),
        "OBJETO": extract_objeto(text),
        "TIPO DE CONTRATO": extract_tipo_contrato(text),
        "PLAZO DE EJECUCIÓN": extract_plazo(text),
        "VALOR DEL CONTRATO": extract_valor(text),
        "FORMA DE PAGO": extract_forma_pago(text),
        "OBLIGACIONES ESPECÍFICAS DEL CONTRATISTA": extract_obligaciones(text),
        "texto_extraido_chars": len(text),
        "extraccion_ok": False,
    }

    required_fields = [
        "Fecha de Elaboración",
        "OBJETO",
        "TIPO DE CONTRATO",
        "PLAZO DE EJECUCIÓN",
        "VALOR DEL CONTRATO",
        "FORMA DE PAGO",
        "OBLIGACIONES ESPECÍFICAS DEL CONTRATISTA",
    ]
    row["extraccion_ok"] = all(bool(str(row.get(field, "")).strip()) for field in required_fields)
    return row



def process_zip(zip_path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    with zipfile.ZipFile(zip_path, "r") as zf:
        pdf_files = [name for name in zf.namelist() if name.lower().endswith(".pdf")]
        if not pdf_files:
            raise ValueError("El ZIP no contiene archivos PDF.")

        total = len(pdf_files)
        for i, name in enumerate(pdf_files, start=1):
            print(f"Procesando {i}/{total}: {Path(name).name}")
            try:
                rows.append(build_row_from_pdf(zf.read(name), name))
            except Exception as e:
                rows.append(
                    {
                        "archivo_pdf": Path(name).name,
                        "Fecha de Elaboración": "",
                        "OBJETO": "",
                        "TIPO DE CONTRATO": "",
                        "PLAZO DE EJECUCIÓN": "",
                        "VALOR DEL CONTRATO": "",
                        "FORMA DE PAGO": "",
                        "OBLIGACIONES ESPECÍFICAS DEL CONTRATISTA": "",
                        "texto_extraido_chars": 0,
                        "extraccion_ok": False,
                        "error": str(e),
                    }
                )
    return rows



def export_rows_to_xlsx(rows: List[Dict[str, object]], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "estudios_previos"

    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in OUTPUT_COLUMNS])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 28,
        "B": 20,
        "C": 45,
        "D": 28,
        "E": 32,
        "F": 42,
        "G": 60,
        "H": 90,
        "I": 18,
        "J": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx in range(2, ws.max_row + 1):
        text_lengths = [len(str(ws[f"{col}{row_idx}"].value or "")) for col in ["C", "F", "G", "H"]]
        max_len = max(text_lengths) if text_lengths else 0
        ws.row_dimensions[row_idx].height = min(120, max(24, 15 + (max_len // 80) * 12))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)





def merge_with_metadata(output_path: Path, metadata_path: Path) -> None:
    import pandas as pd

    estudios_previos_extraidos = pd.read_excel(output_path)
    Metadatos_SIGA_EstudiosPrevios = pd.read_excel(metadata_path)

    merged = estudios_previos_extraidos.merge(
        Metadatos_SIGA_EstudiosPrevios,
        how="left",
        left_on="archivo_pdf",
        right_on="IMGANX_NOMBRE",
    )

    merged.to_excel(output_path, index=False)


def main() -> None:
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        zip_input = sys.argv[1].strip()
    else:
        zip_input = input("Ingrese la ruta del archivo .zip de entrada: ").strip()

    if not zip_input:
        raise ValueError("Debe ingresar la ruta del archivo ZIP de entrada.")

    zip_path = Path(zip_input.strip('"'))
    if not zip_path.exists():
        raise FileNotFoundError(f"No existe el archivo ZIP: {zip_path}")

    if len(sys.argv) >= 3 and sys.argv[2].strip():
        output_dir_input = sys.argv[2].strip()
    else:
        output_dir_input = input(
            'Ingrese la ruta de la carpeta donde desea guardar el output "estudios_previos_extraidos": '
        ).strip()

    if not output_dir_input:
        raise ValueError("Debe ingresar la ruta de salida para guardar el archivo Excel.")

    output_dir = Path(output_dir_input.strip('"')).expanduser()
    output_path = output_dir / "estudios_previos_extraidos.xlsx"

    rows = process_zip(zip_path)
    export_rows_to_xlsx(rows, output_path)

    metadata_input = input(
        "Ingrese la ruta del archivo con  metadatos básicos del contrato asociados a cada uno de los Estudios Previo relacionados en SIGA"
    ).strip()

    if not metadata_input:
        raise ValueError("Debe ingresar la ruta del archivo con metadatos de SIGA.")

    metadata_path = Path(metadata_input.strip('"')).expanduser()
    if not metadata_path.exists():
        raise FileNotFoundError(f"No existe el archivo de metadatos: {metadata_path}")

    merge_with_metadata(output_path, metadata_path)
    print(f"\nProceso terminado. Excel guardado en: {output_path}")


if __name__ == "__main__":
    main()
